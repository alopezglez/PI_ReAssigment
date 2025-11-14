#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para ajustar registros de facturas de restaurante
Autor: Alexander López González
Fecha: 2025-11-12
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random
import sys
import os

def barra_progreso(iteracion, total, prefijo='', sufijo='', longitud=50):
    """Muestra una barra de progreso en consola sin dependencias externas"""
    porcentaje = 100 * (iteracion / float(total))
    lleno = int(longitud * iteracion // total)
    barra = '█' * lleno + '░' * (longitud - lleno)
    print(f'\r{prefijo} │{barra}│ {porcentaje:.0f}% [{iteracion}/{total}] {sufijo}', end='', flush=True)
    if iteracion == total:
        print()  # Nueva línea al terminar

def limpiar_consola():
    """Limpia la consola según el sistema operativo"""
    os.system('cls' if os.name == 'nt' else 'clear')

def solicitar_archivo():
    """Muestra un menú de archivos disponibles en la carpeta 'archivos/'"""
    # Obtener directorio raíz (padre de 'scripts')
    dir_script = os.path.dirname(os.path.abspath(__file__))
    dir_raiz = os.path.dirname(dir_script)  # Subir un nivel desde 'scripts/'
    carpeta_archivos = os.path.join(dir_raiz, 'archivos')
    
    # Crear carpeta si no existe
    if not os.path.exists(carpeta_archivos):
        os.makedirs(carpeta_archivos)
        print("╔═══════════════════════════════════════════════════════════╗")
        print("║    Carpeta 'archivos' creada                              ║")
        print("╚═══════════════════════════════════════════════════════════╝")
        print(f"\nPor favor, copia tus archivos en:")
        print(f"  {carpeta_archivos}")
        input("\nPresiona ENTER cuando hayas copiado los archivos...")
    
    # Buscar archivos válidos (Excel y TXT)
    extensiones_validas = ['.xls', '.xlsx', '.txt', '.csv']
    archivos_disponibles = []
    
    for archivo in os.listdir(carpeta_archivos):
        ext = os.path.splitext(archivo)[1].lower()
        if ext in extensiones_validas:
            ruta_completa = os.path.join(carpeta_archivos, archivo)
            archivos_disponibles.append((archivo, ruta_completa))
    
    if not archivos_disponibles:
        print("╔═══════════════════════════════════════════════════════════╗")
        print("║     ERROR: No hay archivos en la carpeta 'archivos'       ║")
        print("╚═══════════════════════════════════════════════════════════╝")
        print(f"\nCopia archivos (.xls, .xlsx, .txt, .csv) en:")
        print(f"  {carpeta_archivos}")
        input("\nPresiona ENTER para salir...")
        sys.exit(1)
    
    # Mostrar menú
    print("░░░░▒▒▒▒▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▒▒▒▒░░░░")
    print("░░░░▒▒▒▒▓▓▓              SELECCIONA ARCHIVO PRINCIPAL               ▓▓▓▒▒▒▒░░░░")
    print("░░░░▒▒▒▒▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▒▒▒▒░░░░")
    print()
    
    for i, (nombre, _) in enumerate(archivos_disponibles, 1):
        print(f"  [{i}] {nombre}")
    
    print()
    while True:
        try:
            seleccion = input("Selecciona el número del archivo (o 0 para salir): ").strip()
            num = int(seleccion)
            
            if num == 0:
                print("\n👋 Saliendo...")
                sys.exit(0)
            
            if 1 <= num <= len(archivos_disponibles):
                archivo_seleccionado = archivos_disponibles[num - 1][1]
                print(f"\n✅ Archivo seleccionado: {archivos_disponibles[num - 1][0]}")
                return archivo_seleccionado
            else:
                print(f"❌ Por favor, selecciona un número entre 1 y {len(archivos_disponibles)}")
        except ValueError:
            print("❌ Por favor, ingresa un número válido")

def solicitar_ticket_inicial():
    """Solicita el número de ticket inicial"""
    print("\n\n▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▒▒▒▒░░░░")
    while True:
        try:
            ticket = input("\n¿Cuál será el número del PRIMER TICKET? (ej: 1000): ").strip()
            ticket_num = int(ticket)
            if ticket_num > 0:
                return ticket_num
            else:
                print("❌ Por favor, ingresa un número positivo.")
        except ValueError:
            print("❌ Por favor, ingresa un número válido.")

def obtener_parametros_desde_args():
    """Obtiene los parámetros desde los argumentos de línea de comandos"""
    if len(sys.argv) >= 3:
        archivo = sys.argv[1].strip().strip('"').strip("'")
        try:
            ticket = int(sys.argv[2])
            if ticket > 0 and os.path.exists(archivo):
                return archivo, ticket
        except ValueError:
            pass
    return None, None

def cargar_archivo(ruta):
    """Función principal que detecta y carga el archivo según su extensión"""
    extension = os.path.splitext(ruta)[1].lower()
    
    # Si es archivo de texto
    if extension in ['.txt', '.csv']:
        df = cargar_archivo_texto(ruta)
        
        # Preguntar por archivo de TOTALES desde la carpeta 'archivos'
        dir_script = os.path.dirname(os.path.abspath(__file__))
        dir_raiz = os.path.dirname(dir_script)  # Subir un nivel desde 'scripts/'
        carpeta_archivos = os.path.join(dir_raiz, 'archivos')
        
        # Buscar archivos de totales disponibles
        archivos_totales = []
        for archivo in os.listdir(carpeta_archivos):
            # Buscar archivos que contengan "total" en el nombre o sean .txt/.csv
            nombre_lower = archivo.lower()
            ext = os.path.splitext(archivo)[1].lower()
            if ('total' in nombre_lower or ext in ['.txt', '.csv']) and archivo != os.path.basename(ruta):
                ruta_completa = os.path.join(carpeta_archivos, archivo)
                archivos_totales.append((archivo, ruta_completa))
        
        df_totales = None
        
        if archivos_totales:
            print("\n▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
            print("▓      ¿DESEAS USAR UN ARCHIVO DE TOTALES?           ▓")
            print("▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
            print()
            print("  [0] No usar archivo de TOTALES")
            
            for i, (nombre, _) in enumerate(archivos_totales, 1):
                print(f"  [{i}] {nombre}")
            
            print()
            while True:
                try:
                    seleccion = input("Selecciona una opción: ").strip()
                    num = int(seleccion)
                    
                    if num == 0:
                        print("⚠️  No se cargará archivo de TOTALES")
                        break
                    
                    if 1 <= num <= len(archivos_totales):
                        ruta_totales = archivos_totales[num - 1][1]
                        ext_totales = os.path.splitext(ruta_totales)[1].lower()
                        
                        if ext_totales == '.txt':
                            df_totales = cargar_totales_texto(ruta_totales)
                        elif ext_totales in ['.xls', '.xlsx']:
                            df_totales = pd.read_excel(ruta_totales, header=None, engine='openpyxl')
                        else:
                            print("⚠️  Formato no soportado para TOTALES")
                        
                        print(f"✅ Archivo TOTALES cargado: {archivos_totales[num - 1][0]}")
                        break
                    else:
                        print(f"❌ Por favor, selecciona un número entre 0 y {len(archivos_totales)}")
                except ValueError:
                    print("❌ Por favor, ingresa un número válido")
        else:
            print("\n⚠️  No se encontró archivo de TOTALES en la carpeta")
        
        return df, df_totales, ruta, True  # True = es texto
    
    # Si es archivo Excel
    elif extension in ['.xls', '.xlsx']:
        return cargar_excel(ruta)
    
    else:
        print(f"❌ ERROR: Formato no soportado: {extension}")
        print("   Formatos válidos: .xls, .xlsx, .txt, .csv")
        input("\nPresiona ENTER para salir...")
        sys.exit(1)

def convertir_decimales_europeos(df, columnas):
    """Convierte decimales europeos (,) a formato estándar (.)"""
    for col in columnas:
        if col in df.columns:
            # Convertir a string, reemplazar coma por punto, convertir a float
            df[col] = df[col].astype(str).str.replace(',', '.').str.replace(' ', '')
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def cargar_archivo_texto(ruta):
    """Carga archivo de texto delimitado por │"""
    print("\n📂 Cargando archivo de texto...")
    print("   📄 Detectado formato TXT delimitado")
    
    # Leer todas las líneas
    with open(ruta, 'r', encoding='utf-8') as f:
        lineas = f.readlines()
    
    # Filtrar líneas decorativas (que empiezan con ┌, ├, └, │TOTAL, │COCOA SUBTOTAL)
    lineas_validas = []
    encabezado = None
    
    for linea in lineas:
        linea = linea.strip()
        # Ignorar líneas decorativas
        if linea.startswith(('┌', '├', '└', '─')):
            continue
        # Ignorar líneas vacías
        if not linea or linea == '│':
            continue
        # Capturar encabezado
        if 'NºFactura' in linea and encabezado is None:
            encabezado = linea
            continue
        # Ignorar SUBTOTAL y TOTAL
        if 'SUBTOTAL' in linea or linea.startswith('│TOTAL'):
            continue
        # Línea de datos válida
        if linea.startswith('│'):
            lineas_validas.append(linea)
    
    # Parsear encabezado
    if encabezado:
        columnas = [col.strip() for col in encabezado.split('│') if col.strip()]
    else:
        # Encabezado por defecto si no se encuentra
        columnas = ['Local', 'DIA', 'Hora', 'NºFactura', 'Mesa', 'PAX', 'Camarero', 
                   'Base', 'IVA', 'TOTAL Fra', 'Ticket MIX', 'COBRADO', 'TARJETA', 
                   'Descuentos', 'Invitaciones']
    
    # Parsear datos
    datos = []
    for linea in lineas_validas:
        valores = [val.strip() for val in linea.split('│') if val or val == '']
        # Eliminar primer y último elemento si están vacíos (por el │ inicial/final)
        if len(valores) > 0 and valores[0] == '':
            valores = valores[1:]
        if len(valores) > 0 and valores[-1] == '':
            valores = valores[:-1]
        
        if len(valores) > 0:
            datos.append(valores)
    
    # Crear DataFrame
    df = pd.DataFrame(datos, columns=columnas[:len(datos[0])] if datos else columnas)
    
    # Convertir decimales europeos a estándar
    columnas_numericas = ['Base', 'IVA', 'TOTAL Fra', 'COBRADO', 'TARJETA', 'Descuentos', 'Invitaciones']
    df = convertir_decimales_europeos(df, columnas_numericas)
    
    # NORMALIZAR NOMBRES DE COLUMNAS para que coincidan con el formato Excel
    # Esto permite que el resto del código funcione sin cambios
    mapeo_nombres = {
        'Local': 'Local',
        'DIA': 'DIA',
        'Hora': 'Hora',
        'NºFactura': 'NºFactura',
        'Mesa': '_Mesa',  # Columna extra, prefijo _ para ignorar
        'PAX': '_PAX',
        'Camarero': '_Camarero',
        'Base': 'BASE',  # Normalizar a mayúsculas
        'IVA': 'IVA',
        'TOTAL Fra': 'TOTAL Fra',
        'Ticket MIX': '_Ticket MIX',
        'COBRADO': 'COBRADO',
        'TARJETA': 'TARJETA',
        'Descuentos': 'Descuentos',
        'Invitaciones': 'INVITACIONES'  # Normalizar a mayúsculas
    }
    
    # Renombrar columnas
    df = df.rename(columns=mapeo_nombres)
    
    # Reordenar columnas para que coincidan con el formato Excel esperado
    # Orden: Local, DIA, Hora, NºFactura, BASE, IVA, TOTAL Fra, COBRADO, TARJETA, Descuentos, INVITACIONES
    columnas_ordenadas = ['Local', 'DIA', 'Hora', 'NºFactura', 'BASE', 'IVA', 'TOTAL Fra', 
                         'COBRADO', 'TARJETA', 'Descuentos', 'INVITACIONES']
    
    # Mantener solo columnas que existen
    columnas_finales = [c for c in columnas_ordenadas if c in df.columns]
    
    # Agregar columnas extra al final (las que empiezan con _)
    columnas_extra = [c for c in df.columns if c.startswith('_')]
    df = df[columnas_finales + columnas_extra]
    
    print(f"✅ Archivo cargado exitosamente: {len(df)} registros encontrados")
    print(f"   Columnas normalizadas para compatibilidad con formato Excel")
    return df

def cargar_totales_texto(ruta):
    """Carga archivo de totales en formato texto (fecha\\ttotal)"""
    print("\n📂 Cargando archivo TOTALES TXT...")
    
    try:
        # Leer con pandas: delimitador tabulador, sin encabezado
        df_totales = pd.read_csv(ruta, sep='\t', header=None, encoding='utf-8')
        
        # Convertir decimales europeos en la columna de totales
        if len(df_totales.columns) >= 2:
            df_totales[1] = df_totales[1].astype(str).str.replace('.', '').str.replace(',', '.')
            df_totales[1] = pd.to_numeric(df_totales[1], errors='coerce')
        
        print(f"✅ Archivo TOTALES cargado: {len(df_totales)} registros")
        return df_totales
    except Exception as e:
        print(f"❌ Error al cargar TOTALES TXT: {str(e)}")
        return None

def cargar_excel(ruta):
    """Carga el archivo Excel"""
    print("\n📂 Cargando archivo Excel...")
    
    # Determinar el engine según la extensión
    extension = os.path.splitext(ruta)[1].lower()
    
    # Definir engines según el tipo de archivo
    if extension == '.xls':
        engines = ['xlrd', 'openpyxl']
        print("   📄 Detectado formato Excel 2003 (.xls)")
    elif extension == '.xlsx':
        engines = ['openpyxl', 'xlrd']
        print("   📄 Detectado formato Excel moderno (.xlsx)")
    else:
        engines = ['openpyxl', 'xlrd']
        print(f"   📄 Formato detectado: {extension}")
    
    df = None
    df_totales = None
    es_archivo_texto = False
    
    for engine in engines:
        try:
            print(f"   🔄 Intentando cargar con engine='{engine}'...")
            
            # Cargar hoja de registros
            # Intentar primero con la hoja 'Registros', si no existe usar la primera
            try:
                df = pd.read_excel(ruta, sheet_name='Registros', engine=engine)
                print(f"   ✅ Hoja 'Registros' encontrada y cargada")
            except:
                print(f"   ⚠️  Hoja 'Registros' no encontrada, usando la primera hoja")
                df = pd.read_excel(ruta, sheet_name=0, engine=engine)
            
            # Cargar hoja de totales SIN encabezados (header=None)
            try:
                df_totales = pd.read_excel(ruta, sheet_name='TOTALES', engine=engine, header=None)
                print(f"   ✅ Hoja 'TOTALES' encontrada (sin encabezados)")
            except:
                print("   ⚠️  ADVERTENCIA: No se encontró la hoja 'TOTALES'")
                df_totales = None
            
            print(f"✅ Archivo cargado exitosamente: {len(df)} registros encontrados")
            print(f"   Engine utilizado: {engine}")
            return df, df_totales, ruta, False  # False = no es texto
            
        except Exception as e:
            print(f"   ❌ Falló con '{engine}': {str(e)[:100]}")
            if engine == engines[-1]:  # Último intento
                print(f"\n❌ ERROR: No se pudo cargar el archivo con ningún método")
                print(f"\n💡 POSIBLES SOLUCIONES:")
                print("   1. Abre el archivo en Excel y guárdalo como .xlsx (Excel Workbook)")
                print("   2. Verifica que el archivo no esté corrupto")
                print("   3. Cierra el archivo si lo tienes abierto en Excel")
                print("   4. Verifica que el archivo tenga datos en la primera hoja")
                input("\nPresiona ENTER para salir...")
                sys.exit(1)
            else:
                continue  # Intentar con el siguiente engine

def validar_estructura_archivo(df, df_totales, ruta, es_texto=False):
    """Valida que el archivo cumple con los requisitos mínimos"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  VALIDACIÓN PREVIA: Estructura del archivo                                  ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    errores_criticos = []
    advertencias = []
    
    # Validación 1: Verificar datos cargados
    tipo_archivo = "archivo TXT" if es_texto else "archivo Excel"
    print(f"\n🔍 Verificando datos del {tipo_archivo}...")
    if df is None:
        errores_criticos.append("No se pudo cargar el archivo de registros")
    else:
        print("   ✅ Datos de registros cargados correctamente")
    
    if df_totales is None:
        advertencias.append("Archivo de TOTALES no encontrado - No se ajustarán totales")
        print("   ⚠️  Archivo de TOTALES no proporcionado")
    else:
        print("   ✅ Archivo de TOTALES cargado")
        # Validar estructura de TOTALES
        if len(df_totales) < 1:
            advertencias.append("Archivo de TOTALES está vacío")
            print("      ⚠️  El archivo TOTALES está vacío")
        elif len(df_totales.columns) < 2:
            errores_criticos.append("Archivo de TOTALES debe tener al menos 2 columnas (Fecha | Total)")
            print("      ❌ Faltan columnas en archivo TOTALES")
        else:
            print(f"      ✅ {len(df_totales)} registros en archivo TOTALES")
    
    # Validación 2: Verificar columnas mínimas
    print("\n🔍 Verificando columnas requeridas...")
    columnas = df.columns.tolist()
    
    # Columnas esperadas (archivos de texto pueden tener más columnas)
    if es_texto:
        columnas_esperadas = [
            "Local", "DIA", "Hora", "NºFactura", "Mesa", "PAX", "Camarero",
            "Base", "IVA", "TOTAL Fra", "Ticket MIX", "COBRADO", "TARJETA", 
            "Descuentos", "Invitaciones"
        ]
    else:
        columnas_esperadas = [
            "Local", "DIA", "Hora", "NºFactura", "BASE", "IVA", "TOTAL Fra", 
            "COBRADO", "TARJETA", "Descuentos", "INVITACIONES"
        ]
    
    print(f"   Columnas encontradas: {len(columnas)}")
    print(f"   Columnas esperadas: {len(columnas_esperadas)} mínimo")
    
    if len(columnas) < 8:
        errores_criticos.append(f"Faltan columnas. Encontradas: {len(columnas)}, Mínimo requerido: 8")
        print(f"   ❌ Insuficientes columnas (mínimo 8)")
    else:
        print(f"   ✅ Número de columnas correcto")
        
        # Mostrar columnas encontradas vs esperadas
        print("\n   📋 Mapeo de columnas:")
        for i in range(min(len(columnas), len(columnas_esperadas))):
            esperada = columnas_esperadas[i] if i < len(columnas_esperadas) else "N/A"
            encontrada = columnas[i]
            indicador = "✓" if i < 8 else "○"
            print(f"      {indicador} {chr(65+i)}: {encontrada}")
    
    # Validación 3: Verificar número mínimo de registros
    print("\n🔍 Verificando registros...")
    num_registros = len(df)
    print(f"   Registros totales: {num_registros}")
    
    if num_registros < 10:
        errores_criticos.append(f"Muy pocos registros: {num_registros} (mínimo: 10)")
        print(f"   ❌ Muy pocos registros (mínimo 10)")
    elif num_registros < 100:
        advertencias.append(f"Pocos registros: {num_registros}. Verifica que sea correcto")
        print(f"   ⚠️  Pocos registros. ¿Es correcto?")
    else:
        print(f"   ✅ Cantidad de registros adecuada")
    
    # Validación 4: Verificar datos en columnas críticas
    print("\n🔍 Verificando datos en columnas críticas...")
    
    if len(columnas) >= 4:
        # Verificar columna de NºFactura
        col_ticket = columnas[3]
        tickets_vacios = df[col_ticket].isna().sum()
        if tickets_vacios > num_registros * 0.5:
            errores_criticos.append(f"Columna '{col_ticket}' tiene muchos valores vacíos ({tickets_vacios})")
            print(f"   ❌ Columna NºFactura con datos insuficientes")
        else:
            print(f"   ✅ Columna NºFactura con datos válidos")
    
    if len(columnas) >= 7:
        # Verificar columna de TOTAL
        col_total = columnas[6]
        totales_vacios = df[col_total].isna().sum()
        if totales_vacios > num_registros * 0.5:
            errores_criticos.append(f"Columna '{col_total}' tiene muchos valores vacíos ({totales_vacios})")
            print(f"   ❌ Columna TOTAL con datos insuficientes")
        else:
            print(f"   ✅ Columna TOTAL con datos válidos")
    
    # Mostrar resumen
    print("\n▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
    print("▓            RESULTADO DE LA VALIDACIÓN              ▓")
    print("▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓\n")
    
    if errores_criticos:
        print(f"  ❌ ERRORES CRÍTICOS: {len(errores_criticos)}")
        for error in errores_criticos:
            print(f"     • {error}")
        print("\n  ⛔ El archivo NO cumple los requisitos mínimos")
        print("     Por favor, revisa el archivo y vuelve a intentarlo")
        input("\n  Presiona ENTER para salir...")
        sys.exit(1)
    
    if advertencias:
        print(f"  ⚠️  ADVERTENCIAS: {len(advertencias)}")
        for adv in advertencias:
            print(f"     • {adv}")
    
    if not errores_criticos:
        print(f"  ✅ El archivo cumple los requisitos mínimos")
        if not advertencias:
            print("     ✓ Todas las hojas presentes")
            print("     ✓ Columnas correctas")
            print("     ✓ Número de registros adecuado")
    
    print()

def paso1_fusionar_columnas(df):
    """Fusiona columnas H (COBRADO) e I (columna siguiente a COBRADO, típicamente TARJETA o vacía)"""
    print("\n╔═══════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 1: Fusionando columnas COBRADO e INV. TARJETA                         ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    # Mostrar las columnas encontradas
    columnas = df.columns.tolist()
    print(f"\n📋 Columnas encontradas en el archivo:")
    for i, col in enumerate(columnas):
        if i < 15:  # Mostrar solo las primeras 15
            print(f"   {chr(65+i) if i < 26 else i} (índice {i}): {col}")
    
    # Verificar que hay suficientes columnas
    if len(columnas) < 9:
        print(f"\n⚠️  ADVERTENCIA: Solo se encontraron {len(columnas)} columnas")
        print("   Se esperaban al menos 9 columnas")
        if len(columnas) < 8:
            print("   No se puede fusionar. Se omite este paso.")
            return df
    
    # Trabajar con índices de columna (más seguro)
    col_h_idx = 7  # Columna H (COBRADO)
    col_i_idx = 8  # Columna I (siguiente)
    
    col_h = columnas[col_h_idx]
    col_i = columnas[col_i_idx] if col_i_idx < len(columnas) else None
    
    if col_i:
        # Convertir a numérico, forzando errores a 0
        df[col_h] = pd.to_numeric(df[col_h], errors='coerce').fillna(0)
        df[col_i] = pd.to_numeric(df[col_i], errors='coerce').fillna(0)
        
        # Fusionar: H = H + I
        df[col_h] = df[col_h] + df[col_i]
        df[col_i] = ''  # Vaciar columna I
        
        print(f"\n✅ Columnas fusionadas:")
        print(f"   '{col_h}' (H) ahora contiene la suma de H + I")
        print(f"   '{col_i}' (I) vaciada")
    else:
        print(f"\n⚠️  Solo se encontró la columna H. Se omite la fusión.")
    
    return df

def paso2_eliminar_filas_problematicas(df):
    """Elimina filas con subtotales y registros negativos"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 2: Eliminando filas problemáticas                                     ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    inicial = len(df)
    columnas = df.columns.tolist()
    
    # Usar índices de columna
    col_d = columnas[3] if len(columnas) > 3 else None  # NºFactura
    col_e = columnas[4] if len(columnas) > 4 else None  # BASE
    col_f = columnas[5] if len(columnas) > 5 else None  # IVA
    col_g = columnas[6] if len(columnas) > 6 else None  # TOTAL Fra
    col_h = columnas[7] if len(columnas) > 7 else None  # COBRADO
    
    # Eliminar filas con SUBTOTAL
    if col_d:
        df = df[~df[col_d].astype(str).str.contains('SUBTOTAL', case=False, na=False)].copy()
        subtotales = inicial - len(df)
        print(f"✅ Filas de SUBTOTAL eliminadas: {subtotales}")
    else:
        print("⚠️  No se encontró columna de NºFactura")
        subtotales = 0
        df = df.copy()
    
    # Eliminar filas donde haya valores negativos en E, F, G o H
    inicial_neg = len(df)
    
    # Convertir columnas a numérico primero
    columnas_numericas = []
    if col_e:
        df.loc[:, col_e] = pd.to_numeric(df[col_e], errors='coerce')
        columnas_numericas.append(col_e)
    if col_f:
        df.loc[:, col_f] = pd.to_numeric(df[col_f], errors='coerce')
        columnas_numericas.append(col_f)
    if col_g:
        df.loc[:, col_g] = pd.to_numeric(df[col_g], errors='coerce')
        columnas_numericas.append(col_g)
    if col_h:
        df.loc[:, col_h] = pd.to_numeric(df[col_h], errors='coerce')
        columnas_numericas.append(col_h)
    
    # Crear condiciones (usar >= 0 directamente sin fillna para evitar warnings)
    condiciones = []
    for col in columnas_numericas:
        # Comparar con 0, los NaN se convierten automáticamente a False
        condiciones.append((df[col] >= 0) | df[col].isna())
    
    if condiciones:
        from functools import reduce
        import operator
        mascara_final = reduce(operator.and_, condiciones)
        df = df[mascara_final].copy()
        negativos = inicial_neg - len(df)
        print(f"✅ Filas con valores negativos eliminadas: {negativos}")
    else:
        print("⚠️  No se pudieron verificar valores negativos")
        negativos = 0
    
    # Resetear índice
    df = df.reset_index(drop=True)
    
    print(f"✅ Registros restantes: {len(df)}")
    
    return df

def paso6_hacer_tickets_correlativos(df, ticket_inicial):
    """Hace que los tickets sean correlativos por día - SE EJECUTA AL FINAL"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 6: Haciendo tickets correlativos (FINAL)                              ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    columnas = df.columns.tolist()
    col_b = columnas[1]  # DIA
    col_d = columnas[3]  # NºFactura
    
    # Agrupar por día - normalizar a Timestamp sin hora
    df[col_b] = pd.to_datetime(df[col_b], dayfirst=True, errors='coerce').dt.normalize()
    
    # Usar directamente la columna normalizada (sin crear date)
    dias_unicos = sorted(df[col_b].dropna().unique())
    
    print(f"📅 Días encontrados: {len(dias_unicos)}")
    if dias_unicos:
        print(f"   Primer día: {dias_unicos[0]}")
        print(f"   Último día: {dias_unicos[-1]}")
        
        # Analizar meses y días faltantes
        from collections import defaultdict
        import calendar
        
        meses_con_datos = defaultdict(set)
        anos_con_datos = set()
        
        for dia in dias_unicos:
            anos_con_datos.add(dia.year)
            meses_con_datos[dia.year].add(dia.month)
        
        # Mostrar análisis por año
        for ano in sorted(anos_con_datos):
            meses = sorted(meses_con_datos[ano])
            nombres_meses = [calendar.month_name[m] for m in meses]
            
            if len(meses) == 12:
                print(f"   📊 Analizado año {ano}: Completo (12 meses)")
            else:
                meses_str = ", ".join(nombres_meses)
                print(f"   📊 Analizado año {ano}: {meses_str}")
    
    ticket_actual = ticket_inicial
    
    for dia in dias_unicos:
        mask = df[col_b] == dia
        indices = df[mask].index.tolist()
        num_tickets = len(indices)
        
        for idx in indices:
            df.at[idx, col_d] = ticket_actual
            ticket_actual += 1
    
    print(f"✅ Tickets renumerados desde {ticket_inicial} hasta {ticket_actual - 1}")
    print(f"   Total de días procesados: {len(dias_unicos)}")
    
    return df, ticket_actual - 1

def paso3_ajustar_totales_por_dia(df, df_totales):
    """Ajusta los totales de cada día para que coincidan con la hoja TOTALES"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 3: Ajustando totales por día                                          ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    columnas = df.columns.tolist()
    col_a = columnas[0]  # Local
    col_b = columnas[1]  # DIA
    col_c = columnas[2]  # Hora
    col_d = columnas[3]  # NºFactura
    col_e = columnas[4]  # BASE
    col_f = columnas[5]  # IVA
    col_g = columnas[6]  # TOTAL Fra
    col_h = columnas[7]  # COBRADO
    
    # Crear diccionario de totales esperados
    totales_esperados = {}
    if df_totales is not None:
        print(f"\n📋 Analizando hoja TOTALES:")
        print(f"   Filas encontradas: {len(df_totales)}")
        print(f"   Primeras 3 filas como muestra:")
        
        # Mostrar muestra de datos
        for i in range(min(3, len(df_totales))):
            print(f"      Fila {i+1}: Fecha={df_totales.iloc[i, 0]} | Total={df_totales.iloc[i, 1]}")
        
        # Leer datos: Columna 0 = Fecha, Columna 1 = Total
        for idx, row in df_totales.iterrows():
            fecha = pd.to_datetime(row[0], dayfirst=True, errors='coerce')
            if pd.notna(fecha):
                total = float(row[1]) if pd.notna(row[1]) else None
                if total is not None:
                    totales_esperados[fecha.normalize()] = total
        
        print(f"   Total de fechas válidas procesadas: {len(totales_esperados)}")
        if totales_esperados:
            primer_total = min(totales_esperados.keys())
            ultimo_total = max(totales_esperados.keys())
            print(f"   Rango: {primer_total} a {ultimo_total}")
    else:
        print("⚠️  No se encontró la hoja TOTALES")
    
    if not totales_esperados:
        print("⚠️  ADVERTENCIA: No hay totales en la hoja TOTALES")
        print("   Se mantendrán los totales calculados de los registros")
        print("   Para ajustar totales, verifica que:")
        print("   - La hoja se llame exactamente 'TOTALES'")
        print("   - Columna A tenga fechas en formato dd/mm/yyyy")
        print("   - Columna B tenga los totales esperados")
        return df
    
    # Normalizar fechas a Timestamp sin hora
    if '_fecha_temp' in df.columns:
        df = df.drop('_fecha_temp', axis=1)
    
    if not pd.api.types.is_datetime64_any_dtype(df[col_b]):
        df[col_b] = pd.to_datetime(df[col_b], dayfirst=True, errors='coerce')
    
    # Normalizar para eliminar hora (mantener como Timestamp)
    df[col_b] = df[col_b].dt.normalize()
    df['_fecha_temp'] = df[col_b]
    
    # Procesar cada día
    dias_unicos = sorted([d for d in df['_fecha_temp'].dropna().unique()])
    filas_nuevas = []
    
    dias_sin_total = []
    dias_ajustados = []
    dias_correctos = []
    
    print(f"\n🔍 Procesando {len(dias_unicos)} días...")
    for idx, dia in enumerate(dias_unicos, 1):
        barra_progreso(idx, len(dias_unicos), prefijo='PASO 3: Ajustando totales', sufijo='días')
        mask = df['_fecha_temp'] == dia
        registros_dia = df[mask].copy()
        
        total_actual = registros_dia[col_g].sum()
        total_esperado = totales_esperados.get(dia)
        
        if total_esperado is None:
            dias_sin_total.append(dia)
            continue
        
        diferencia = total_esperado - total_actual
        
        if abs(diferencia) < 0.01:
            dias_correctos.append(dia)
            continue
        
        dias_ajustados.append(dia)
        
        # Obtener datos del último registro del día
        ultimo_registro = registros_dia.iloc[-1]
        ultimo_ticket = int(ultimo_registro[col_d])
        hora_base = ultimo_registro[col_c]
        local = ultimo_registro[col_a]
        
        # Crear tickets de ajuste
        num_tickets = max(1, int(abs(diferencia) / 50))  # Dividir en tickets de ~50€
        ajuste_por_ticket = diferencia / num_tickets
        
        for i in range(num_tickets):
            # Calcular hora (incrementar minutos)
            if isinstance(hora_base, str):
                try:
                    hora_dt = datetime.strptime(hora_base, '%H:%M:%S')
                except:
                    try:
                        hora_dt = datetime.strptime(hora_base, '%H:%M')
                    except:
                        hora_dt = datetime.strptime('23:59:00', '%H:%M:%S')
            else:
                hora_dt = hora_base
            
            nueva_hora = (hora_dt + timedelta(minutes=i+1)).strftime('%H:%M')
            
            # Calcular BASE e IVA (IVA = 7%) - Redondear a 2 decimales
            total_ticket = round(ajuste_por_ticket, 2)
            base = round(total_ticket / 1.07, 2)
            iva = round(total_ticket - base, 2)
            
            nueva_fila = {
                col_a: local,
                col_b: dia,  # Ya es Timestamp normalizado
                col_c: nueva_hora,
                col_d: ultimo_ticket + i + 1,
                col_e: round(base, 2),
                col_f: round(iva, 2),
                col_g: round(total_ticket, 2),
                col_h: round(total_ticket, 2),
            }
            
            # Añadir columnas restantes vacías
            for col_idx in range(8, len(columnas)):
                nueva_fila[columnas[col_idx]] = ''
            
            filas_nuevas.append(nueva_fila)
    
    # Añadir las filas nuevas al DataFrame
    if filas_nuevas:
        df_nuevas = pd.DataFrame(filas_nuevas)
        df = pd.concat([df, df_nuevas], ignore_index=True)
        
        # Reordenar por día y hora
        df = df.sort_values([col_b, col_c]).reset_index(drop=True)
        
        print(f"\n✅ Se agregaron {len(filas_nuevas)} tickets de ajuste")
    
    # Mostrar resumen
    print(f"\n📊 RESUMEN DEL PASO 4:")
    print(f"   ✅ Días correctos: {len(dias_correctos)}")
    print(f"   🔧 Días ajustados: {len(dias_ajustados)}")
    if dias_sin_total:
        print(f"   ⚠️  Días SIN TOTAL en hoja TOTALES: {len(dias_sin_total)}")
        print(f"      ACCIÓN REQUERIDA: Verifica la hoja TOTALES")
        print(f"      Días afectados: {dias_sin_total[0]} a {dias_sin_total[-1]}")
    
    return df

def paso4_añadir_ticket_negativo(df):
    """NO HACE NADA - El ticket de compensación se añade en paso5 según la diferencia"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 4: Preparando para añadir tickets de compensación                     ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    print("⏩ Este paso se ejecutará en el PASO 5 junto con los tickets intermedios")
    
    # Solo normalizar fechas para el siguiente paso
    columnas = df.columns.tolist()
    col_b = columnas[1]  # DIA
    
    if '_fecha_temp' in df.columns:
        df = df.drop('_fecha_temp', axis=1)
    
    if not pd.api.types.is_datetime64_any_dtype(df[col_b]):
        df[col_b] = pd.to_datetime(df[col_b], dayfirst=True, errors='coerce')
    
    df[col_b] = df[col_b].dt.normalize()
    df['_fecha_temp'] = df[col_b]
    
    return df

def paso5_compensar_tickets_negativos(df, df_totales):
    """Añade tickets de compensación según la diferencia y ticket final de invitación"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 5: Aplicando lógica de compensación y tickets de invitación           ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    columnas = df.columns.tolist()
    col_a = columnas[0]  # Local
    col_b = columnas[1]  # DIA
    col_c = columnas[2]  # Hora
    col_d = columnas[3]  # NºFactura
    col_e = columnas[4]  # BASE
    col_f = columnas[5]  # IVA
    col_g = columnas[6]  # TOTAL Fra
    col_h = columnas[7]  # COBRADO
    col_j = columnas[9] if len(columnas) > 9 else 'INVITACIONES'  # INVITACIONES
    
    # Crear diccionario de totales esperados
    totales_esperados = {}
    if df_totales is not None:
        for _, row in df_totales.iterrows():
            fecha = pd.to_datetime(row[0], dayfirst=True, errors='coerce')
            if pd.notna(fecha):
                total = float(row[1]) if pd.notna(row[1]) else None
                if total is not None:
                    totales_esperados[fecha.normalize()] = total
    
    # Normalizar fechas a Timestamp sin hora
    if '_fecha_temp' in df.columns:
        df = df.drop('_fecha_temp', axis=1)
    
    if not pd.api.types.is_datetime64_any_dtype(df[col_b]):
        df[col_b] = pd.to_datetime(df[col_b], dayfirst=True, errors='coerce')
    
    df[col_b] = df[col_b].dt.normalize()
    df['_fecha_temp'] = df[col_b]
    dias_unicos = sorted([d for d in df['_fecha_temp'].dropna().unique()])
    
    # Procesar cada día
    df_final = pd.DataFrame()
    
    print(f"\n🔍 Procesando {len(dias_unicos)} días...")
    dias_caso1 = 0  # Total < Esperado
    dias_caso2 = 0  # Total > Esperado
    dias_caso3 = 0  # Total = Esperado
    
    for idx, dia in enumerate(dias_unicos, 1):
        barra_progreso(idx, len(dias_unicos), prefijo='PASO 5: Compensación', sufijo='días')
        mask = df['_fecha_temp'] == dia
        registros_dia = df[mask].copy().reset_index(drop=True)
        
        # Calcular totales
        total_actual = registros_dia[col_g].sum()
        total_esperado = totales_esperados.get(dia, total_actual)
        diferencia = total_esperado - total_actual
        
        # Obtener datos de referencia
        local = registros_dia.iloc[0][col_a]
        ultimo_registro = registros_dia.iloc[-1]
        
        # Calcular importe aleatorio (0.8% - 1.0% del total esperado)
        porcentaje_aleatorio = random.uniform(0.008, 0.010)
        importe_aleatorio = round(total_esperado * porcentaje_aleatorio, 2)
        
        # CASO 1: Total < Esperado (Faltan ingresos)
        if diferencia > 0.01:
            dias_caso1 += 1
            
            # Total a compensar con tickets intermedios: diferencia + importe_aleatorio
            total_compensacion = round(diferencia + importe_aleatorio, 2)
            
            # Dividir en 2 tickets con importes aleatorios
            porcentaje_ticket1 = random.uniform(0.4, 0.6)
            importe_ticket1 = round(total_compensacion * porcentaje_ticket1, 2)
            importe_ticket2 = round(total_compensacion - importe_ticket1, 2)
            
            # Ticket final de invitación NEGATIVO
            ticket_invitacion_importe = -importe_aleatorio
            
        # CASO 2: Total > Esperado (Sobran ingresos)
        elif diferencia < -0.01:
            dias_caso2 += 1
            
            # NO crear tickets intermedios
            importe_ticket1 = 0
            importe_ticket2 = 0
            
            # Ticket final de invitación NEGATIVO por la diferencia
            ticket_invitacion_importe = diferencia  # Ya es negativo
            
        # CASO 3: Total = Esperado (Exacto)
        else:
            dias_caso3 += 1
            
            # Total a compensar con tickets intermedios: importe_aleatorio
            total_compensacion = importe_aleatorio
            
            # Dividir en 2 tickets con importes aleatorios
            porcentaje_ticket1 = random.uniform(0.4, 0.6)
            importe_ticket1 = round(total_compensacion * porcentaje_ticket1, 2)
            importe_ticket2 = round(total_compensacion - importe_ticket1, 2)
            
            # Ticket final de invitación NEGATIVO
            ticket_invitacion_importe = -importe_aleatorio
        
        # INSERTAR TICKETS INTERMEDIOS (si corresponde)
        registros_finales = registros_dia.copy()
        
        if importe_ticket1 > 0 and importe_ticket2 > 0:
            # Elegir posiciones aleatorias ENTRE los tickets normales
            num_registros = len(registros_dia)
            if num_registros < 2:
                posiciones = [0, 0]
            else:
                # Insertar en el medio del día (entre 30% y 70%)
                pos_min = int(num_registros * 0.3)
                pos_max = int(num_registros * 0.7)
                if pos_max <= pos_min:
                    pos_max = pos_min + 1
                
                pos1 = random.randint(pos_min, pos_max)
                pos2 = random.randint(pos_min, pos_max)
                if pos2 == pos1:
                    pos2 = min(pos1 + random.randint(1, 3), num_registros - 1)
                
                posiciones = sorted([pos1, pos2])
            
            # Crear los 2 tickets de compensación
            tickets_compensacion = []
            for pos, importe in zip(posiciones, [importe_ticket1, importe_ticket2]):
                # Obtener hora de referencia
                if pos < len(registros_dia):
                    hora_ref = registros_dia.iloc[pos][col_c]
                else:
                    hora_ref = registros_dia.iloc[-1][col_c]
                
                # Calcular hora (añadir segundos aleatorios)
                if isinstance(hora_ref, str):
                    try:
                        hora_dt = datetime.strptime(hora_ref, '%H:%M:%S')
                    except:
                        try:
                            hora_dt = datetime.strptime(hora_ref, '%H:%M')
                        except:
                            hora_dt = datetime.strptime('12:00:00', '%H:%M:%S')
                else:
                    hora_dt = hora_ref
                
                nueva_hora = (hora_dt + timedelta(seconds=random.randint(10, 50))).strftime('%H:%M')
                
                # Calcular BASE e IVA
                base = round(importe / 1.07, 2)
                iva = round(importe - base, 2)
                
                nueva_fila = {
                    col_a: local,
                    col_b: dia,
                    col_c: nueva_hora,
                    col_d: 999999,  # Temporal
                    col_e: round(base, 2),
                    col_f: round(iva, 2),
                    col_g: round(importe, 2),
                    col_h: round(importe, 2),
                }
                
                # Añadir columnas restantes vacías
                for col_idx in range(8, len(columnas)):
                    nueva_fila[columnas[col_idx]] = ''
                
                tickets_compensacion.append(nueva_fila)
            
            # Insertar tickets en posiciones correspondientes
            for i, (pos, ticket) in enumerate(zip(posiciones, tickets_compensacion)):
                pos_ajustada = pos + i  # Ajustar por tickets ya insertados
                ticket_df = pd.DataFrame([ticket])
                registros_finales = pd.concat([
                    registros_finales.iloc[:pos_ajustada],
                    ticket_df,
                    registros_finales.iloc[pos_ajustada:]
                ], ignore_index=True)
        
        # AÑADIR TICKET FINAL DE INVITACIÓN
        # Calcular BASE e IVA del ticket de invitación
        base_inv = round(ticket_invitacion_importe / 1.07, 2)
        iva_inv = round(ticket_invitacion_importe - base_inv, 2)
        
        ticket_invitacion = {
            col_a: local,
            col_b: dia,
            col_c: "23:59",
            col_d: 999999,  # Temporal
            col_e: round(base_inv, 2),
            col_f: round(iva_inv, 2),
            col_g: round(ticket_invitacion_importe, 2),
            col_h: 0,  # COBRADO = 0
        }
        
        # Añadir columnas restantes vacías PRIMERO
        for col_idx in range(len(columnas)):
            if columnas[col_idx] not in ticket_invitacion:
                ticket_invitacion[columnas[col_idx]] = ''
        
        # LUEGO establecer INVITACIONES con el importe negativo
        ticket_invitacion[col_j] = round(ticket_invitacion_importe, 2)
        
        ticket_inv_df = pd.DataFrame([ticket_invitacion])
        registros_finales = pd.concat([registros_finales, ticket_inv_df], ignore_index=True)
        
        df_final = pd.concat([df_final, registros_finales], ignore_index=True)
    
    # Ordenar por fecha y hora
    df_final = df_final.sort_values([col_b, col_c]).reset_index(drop=True)
    
    print(f"\n✅ Completado:")
    print(f"   • Caso 1 (Total < Esperado): {dias_caso1} días")
    print(f"   • Caso 2 (Total > Esperado): {dias_caso2} días")
    print(f"   • Caso 3 (Total = Esperado): {dias_caso3} días")
    
    return df_final

def paso7_verificacion_final(df, df_totales):
    """Verifica la integridad final de los datos"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  PASO 7: Verificación final de integridad                                   ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    columnas = df.columns.tolist()
    col_b = columnas[1]  # DIA
    col_g = columnas[6]  # TOTAL Fra
    col_j = columnas[9] if len(columnas) > 9 else None  # INVITACIONES (negativos)
    
    # Normalizar fechas a Timestamp sin hora
    if '_fecha_temp' in df.columns:
        df = df.drop('_fecha_temp', axis=1)
    
    if not pd.api.types.is_datetime64_any_dtype(df[col_b]):
        df[col_b] = pd.to_datetime(df[col_b], dayfirst=True, errors='coerce')
    
    df[col_b] = df[col_b].dt.normalize()
    df['_fecha_temp'] = df[col_b]
    dias_unicos = sorted([d for d in df['_fecha_temp'].dropna().unique()])
    
    # Cargar totales esperados
    totales_esperados = {}
    if df_totales is not None:
        for _, row in df_totales.iterrows():
            fecha = pd.to_datetime(row[0], dayfirst=True, errors='coerce')
            if pd.notna(fecha):
                total = float(row[1]) if pd.notna(row[1]) else None
                if total is not None:
                    totales_esperados[fecha.normalize()] = total
    
    errores = []
    advertencias = []
    dias_ok = 0
    dias_con_error = []
    
    print(f"\n🔍 Analizando {len(dias_unicos)} días...")
    for idx, dia in enumerate(dias_unicos, 1):
        barra_progreso(idx, len(dias_unicos), prefijo='PASO 7: Verificación final', sufijo='días')
        mask = df['_fecha_temp'] == dia
        registros_dia = df[mask].copy()
        
        # Verificar 1: Contar tickets negativos (donde TOTAL es negativo)
        negativos = registros_dia[registros_dia[col_g] < 0]
        num_negativos = len(negativos)
        
        # Verificar 2: Total del día
        total_dia = registros_dia[col_g].sum()
        total_esperado = totales_esperados.get(dia)
        
        # Indicadores
        icono_negativos = "✅" if num_negativos == 1 else "❌"
        
        if total_esperado is not None:
            diferencia = abs(total_esperado - total_dia)
        else:
            diferencia = 0
        
        # Registrar errores
        tiene_error = False
        if num_negativos != 1:
            errores.append(f"{dia}: Tiene {num_negativos} tickets negativos (esperado: 1)")
            tiene_error = True
        
        if total_esperado is not None and diferencia >= 0.01:
            errores.append(f"{dia}: Diferencia en total: {diferencia:.2f}€")
            tiene_error = True
        
        if total_esperado is None:
            advertencias.append(f"{dia}: Sin total de referencia en hoja TOTALES")
        
        if tiene_error:
            dias_con_error.append((dia, num_negativos, total_dia, total_esperado, diferencia))
        elif num_negativos == 1 and (total_esperado is None or diferencia < 0.01):
            dias_ok += 1
    
    # Mostrar resumen final
    print("\n" + "═" * 63)
    print("  📊 RESULTADO DE LA VERIFICACIÓN")
    print("═" * 63)
    print(f"  ✅ Días correctos: {dias_ok}/{len(dias_unicos)}")
    print(f"  ❌ Días con errores: {len(set([e.split(':')[0] for e in errores]))}")
    print(f"  ⚠️  Advertencias: {len(advertencias)}")
    
    if dias_con_error:
        print(f"\n  ❌ DÍAS CON ERRORES:")
        print(f"  {'Fecha':<12} {'Negativos':<10} {'Total Real':<12} {'Esperado':<12} {'Diferencia':<12}")
        print(f"  {'-'*60}")
        for dia, negs, real, esperado, dif in dias_con_error[:10]:
            esp_str = f"{esperado:.2f}€" if esperado else "N/A"
            dif_str = f"{dif:+.2f}€" if esperado else "N/A"
            print(f"  {dia}  {negs:<10} {real:>10.2f}€  {esp_str:>10}  {dif_str:>10}")
        if len(dias_con_error) > 10:
            print(f"  ... y {len(dias_con_error) - 10} días con errores más")
    
    if advertencias and not errores:
        print(f"\n  ⚠️  ADVERTENCIAS:")
        for adv in advertencias[:5]:  # Mostrar máximo 5
            print(f"     • {adv}")
        if len(advertencias) > 5:
            print(f"     ... y {len(advertencias) - 5} advertencias más")
    
    if not errores and not advertencias:
        print("\n  🎉 ¡PERFECTO! Todos los datos son correctos")
        print("     ✓ Un ticket negativo por día")
        print("     ✓ Totales coinciden con hoja TOTALES")
        print("     ✓ Tickets correlativos")
    
    return df

def guardar_resultado(df, ruta_original):
    """Guarda el resultado en un nuevo archivo Excel"""
    print("\n╔═════════════════════════════════════════════════════════════════════════════╗")
    print("║  GUARDANDO RESULTADO                                                        ║")
    print("╚═════════════════════════════════════════════════════════════════════════════╝")
    
    # Eliminar columnas temporales antes de guardar
    if '_fecha_temp' in df.columns:
        df = df.drop('_fecha_temp', axis=1)
    
    # ELIMINAR COLUMNAS NO DESEADAS
    columnas_eliminar = ['Descuentos', '_Mesa', '_PAX', '_Camarero', '_Ticket MIX', 'Tarjeta', 'Tarjetas']
    columnas_encontradas = [col for col in columnas_eliminar if col in df.columns]
    
    if columnas_encontradas:
        print(f"\n🗑️  Eliminando columnas: {', '.join(columnas_encontradas)}")
        df = df.drop(columns=columnas_encontradas, errors='ignore')
    
    # Crear carpeta 'resultados' si no existe
    dir_script = os.path.dirname(os.path.abspath(__file__))
    dir_raiz = os.path.dirname(dir_script)  # Subir un nivel desde 'scripts/'
    carpeta_resultados = os.path.join(dir_raiz, 'resultados')
    
    if not os.path.exists(carpeta_resultados):
        os.makedirs(carpeta_resultados)
        print(f"📁 Carpeta 'resultados' creada")
    
    # Crear nombre del archivo de salida
    nombre_original = os.path.basename(ruta_original)
    nombre_sin_ext = os.path.splitext(nombre_original)[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    print(f"\n💾 Preparando archivo de salida...")
    
    try:
        # Obtener nombres de columnas después de eliminar
        columnas = df.columns.tolist()
        col_b = columnas[1] if len(columnas) > 1 else None  # DIA
        col_e = columnas[4] if len(columnas) > 4 else None  # BASE
        col_f = columnas[5] if len(columnas) > 5 else None  # IVA
        col_g = columnas[6] if len(columnas) > 6 else None  # TOTAL Fra
        col_h = columnas[7] if len(columnas) > 7 else None  # COBRADO
        
        # 1. Convertir fechas a solo fecha (sin hora)
        if col_b and pd.api.types.is_datetime64_any_dtype(df[col_b]):
            df[col_b] = df[col_b].dt.date
        
        # 2. Forzar 2 decimales en columnas numéricas
        for col in [col_e, col_f, col_g, col_h]:
            if col:
                df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
        
        # PREGUNTAR FORMATO DE SALIDA
        print("\n╔═══════════════════════════════════════════════════════════╗")
        print("║        ¿EN QUÉ FORMATO DESEAS GUARDAR EL ARCHIVO?         ║")
        print("╚═══════════════════════════════════════════════════════════╝")
        print("\n  [1] Excel (.xlsx) - Formato con estilos y colores")
        print("  [2] Texto (.txt) - Formato de tabla ASCII")
        print()
        
        while True:
            formato = input("Selecciona una opción (1 o 2): ").strip()
            if formato in ['1', '2']:
                break
            print("⚠️  Por favor, selecciona 1 o 2")
        
        if formato == '1':
            # GUARDAR COMO EXCEL
            nombre_salida = f"{nombre_sin_ext}_AJUSTADO_{timestamp}.xlsx"
            ruta_salida = os.path.join(carpeta_resultados, nombre_salida)
            
            print(f"\n💾 Guardando archivo Excel...")
            print(f"⏳ Procesando {len(df)} registros...")
            
            df.to_excel(ruta_salida, index=False, engine='openpyxl')
            
            # 3. Aplicar formato de número con 2 decimales en Excel
            from openpyxl import load_workbook
            from openpyxl.styles import numbers
            
            print("   🎨 Aplicando formato de números...")
            wb = load_workbook(ruta_salida)
            ws = wb.active
            
            # Aplicar formato de número con 2 decimales a columnas numéricas
            for col_idx in range(5, min(9, len(columnas) + 1)):
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        try:
                            float(cell.value)
                            cell.number_format = '0.00'
                        except:
                            pass
            
            print("   💾 Guardando cambios...")
            wb.save(ruta_salida)
            
            print(f"\n✅ Archivo Excel guardado exitosamente")
            print(f"📁 resultados/{nombre_salida}")
        
        else:
            # GUARDAR COMO TEXTO CON FORMATO ASCII
            nombre_salida = f"{nombre_sin_ext}_AJUSTADO_{timestamp}.txt"
            ruta_salida = os.path.join(carpeta_resultados, nombre_salida)
            
            print(f"\n💾 Guardando archivo de texto...")
            print(f"⏳ Procesando {len(df)} registros (puede tardar unos segundos)...")
            
            with open(ruta_salida, 'w', encoding='utf-8-sig') as f:
                # Calcular anchos de columna
                print("   📐 Calculando anchos de columna...")
                anchos = []
                for col in df.columns:
                    max_ancho = max(
                        len(str(col)),
                        df[col].astype(str).str.len().max() if len(df) > 0 else 0
                    )
                    anchos.append(max_ancho + 2)
                
                # Línea superior
                print("   🎨 Generando encabezados...")
                f.write('┌' + '┬'.join(['─' * ancho for ancho in anchos]) + '┐\n')
                
                # Encabezados
                encabezados = []
                for col, ancho in zip(df.columns, anchos):
                    encabezados.append(str(col).ljust(ancho))
                f.write('│' + '│'.join(encabezados) + '│\n')
                
                # Línea separadora
                f.write('├' + '┼'.join(['─' * ancho for ancho in anchos]) + '┤\n')
                
                # Datos con barra de progreso
                print("   📝 Escribiendo datos...")
                total_filas = len(df)
                for idx, (_, row) in enumerate(df.iterrows(), 1):
                    valores = []
                    for val, ancho in zip(row, anchos):
                        # Formatear valores según tipo
                        if pd.isna(val) or val == '' or val == ' ':
                            # Valores vacíos o nulos
                            val_str = ''
                        elif isinstance(val, (int, float)):
                            # Números: formatear con 2 decimales
                            val_str = f"{val:.2f}"
                        else:
                            # Texto: convertir a string
                            val_str = str(val)
                        valores.append(val_str.ljust(ancho))
                    f.write('│' + '│'.join(valores) + '│\n')
                    
                    # Mostrar progreso cada 100 filas
                    if idx % 100 == 0 or idx == total_filas:
                        barra_progreso(idx, total_filas, prefijo='      ', sufijo='filas escritas')
                
                # Línea inferior
                print("\n   🎨 Finalizando formato...")
                f.write('└' + '┴'.join(['─' * ancho for ancho in anchos]) + '┘\n')
            
            print(f"\n✅ Archivo de texto guardado exitosamente")
            print(f"📁 resultados/{nombre_salida}")
            print(f"   ✓ Formato tabla ASCII con caracteres de caja")
        
        print(f"\n📊 Total de registros: {len(df)}")
        print(f"📋 Columnas eliminadas: {len(columnas_encontradas)}")
        print(f"📂 Ubicación: {carpeta_resultados}")
        print(f"   ✓ Fechas sin hora")
        print(f"   ✓ Importes con 2 decimales")
        
        return ruta_salida
    except Exception as e:
        print(f"❌ ERROR al guardar: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def mostrar_resumen_final(df, ruta_salida):
    """Muestra un resumen final del proceso"""
    print("\n░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░")
    print("▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒")
    print("▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
    print("▓                               RESUMEN FINAL                                 ▓")
    print("▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
    print("▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒")
    print("░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░")
    
    columnas = df.columns.tolist()
    col_b = columnas[1]  # DIA
    col_g = columnas[6]  # TOTAL Fra
    
    # Asegurar columna temporal de fecha
    if '_fecha_temp' in df.columns:
        df = df.drop('_fecha_temp', axis=1)
    
    # Asegurarse de que col_b es datetime
    if not pd.api.types.is_datetime64_any_dtype(df[col_b]):
        df[col_b] = pd.to_datetime(df[col_b], dayfirst=True, errors='coerce')
    
    df['_fecha_temp'] = df[col_b].dt.date
    
    dias_unicos = sorted([d for d in df['_fecha_temp'].unique() if d is not None and not pd.isna(d)])
    
    print(f"\n📅 Días procesados: {len(dias_unicos)}")
    print(f"📝 Total de registros: {len(df)}")
    print(f"\n💰 Totales por día:")
    print("─" * 63)
    
    for dia in dias_unicos:
        mask = df['_fecha_temp'] == dia
        registros = df[mask]
        total = registros[col_g].sum()
        num_tickets = len(registros)
        
        print(f"  {dia}: {total:>10.2f}€  ({num_tickets} tickets)")
    
    total_general = df[col_g].sum()

    print("\n░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░")
    print("▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒")
    print("▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
    print(f" 💰 TOTAL GENERAL: {total_general:>10.2f}€")
    print(f" 🎫 TOTAL TICKETS: {len(df)}")
    print("▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓")
    print("▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒▒")
    print("░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░")
        
    print(f"\n✅ Proceso completado exitosamente")
    print(f"📁 Archivo guardado en:\n   {ruta_salida}")

def procesar_archivo(usar_argumentos=False):
    """Procesa un archivo completo"""
    # SIEMPRE usar el menú de selección (no argumentos)
    ruta = solicitar_archivo()
    ticket_inicial = solicitar_ticket_inicial()
    
    # Cargar archivo (detecta automáticamente si es Excel o TXT)
    df, df_totales, ruta_original, es_texto = cargar_archivo(ruta)
    
    # VALIDACIÓN PREVIA: Verificar estructura del archivo
    validar_estructura_archivo(df, df_totales, ruta_original, es_texto)
    
    # PASO 1: Fusionar columnas
    df = paso1_fusionar_columnas(df)
    
    # PASO 2: Eliminar filas problemáticas
    df = paso2_eliminar_filas_problematicas(df)
    
    # PASO 3: Ajustar totales por día
    df = paso3_ajustar_totales_por_dia(df, df_totales)
    
    # PASO 4: Añadir ticket negativo al final de cada día
    df = paso4_añadir_ticket_negativo(df)
    
    # PASO 5: Compensar tickets negativos
    df = paso5_compensar_tickets_negativos(df, df_totales)
    
    # PASO 6: Hacer tickets correlativos (AL FINAL)
    df, ultimo_ticket = paso6_hacer_tickets_correlativos(df, ticket_inicial)
    
    # PASO 7: Verificación final
    df = paso7_verificacion_final(df, df_totales)
    
    # Guardar resultado
    ruta_salida = guardar_resultado(df, ruta_original)
    
    if ruta_salida:
        # Mostrar resumen final
        mostrar_resumen_final(df, ruta_salida)

def main():
    """Función principal con bucle de repetición"""
    try:
        while True:
            # Procesar archivo (SIEMPRE usar menú, nunca argumentos)
            procesar_archivo(usar_argumentos=False)
            
            # Preguntar si quiere procesar otro archivo
            print("\n" + "░" * 60)
            print("\n¿Qué deseas hacer?")
            print("  [1] Procesar otro archivo")
            print("  [2] Salir")
            print()
            
            opcion = input("Selecciona una opción (1 o 2): ").strip()
            
            if opcion == "2":
                print("\n👋 ¡Hasta pronto!")
                break
            elif opcion != "1":
                print("\n⚠️  Opción no válida. Saliendo...")
                break
            
            # Limpiar consola antes de la siguiente iteración
            limpiar_consola()
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso cancelado por el usuario")
        input("\nPresiona ENTER para salir...")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO: {str(e)}")
        import traceback
        traceback.print_exc()
        input("\nPresiona ENTER para salir...")
        sys.exit(1)

if __name__ == "__main__":
    main()
